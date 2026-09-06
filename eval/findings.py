#!/usr/bin/env python3
"""Package ordinary workspace files for the findings judge; no agent preprocessing.

Rendering and text budgets copied from browser-use/new-eval-platform at
ddc48ee93ea863c26d78c45a09760e4951b48a3d, harnesses/bcode/run.py.
The four rendering/budget functions below preserve its evidence semantics.
"""
import io
import json
import sys
from pathlib import Path

OUTPUT_TOTAL_MAX_CHARS = 600_000
OUTPUT_MAX_FILES = 50

def _render_binary(path: Path, raw: bytes) -> tuple[str, str] | None:
    """Best-effort (text, kind) preview for xlsx/pdf deliverables, so the
    judge grades the content instead of the agent's self-report about it.
    These were the two binary formats that mattered on the CMU calibration
    waves (xlsx 8% of runs, pdf 14%). Returns None when extraction fails."""
    suffix = path.suffix.lower()
    try:
        if suffix in (".xlsx", ".xlsm"):
            import csv
            from openpyxl import load_workbook

            # data_only=False: agent-generated workbooks carry no cached
            # formula results, and data_only=True would render every formula
            # cell blank. Literal values render identically either way;
            # formula cells show their expression, which is still evidence.
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=False)
            out = io.StringIO()
            for ws in wb.worksheets:
                out.write(f"[sheet: {ws.title}]\n")
                writer = csv.writer(out)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if v is None else v for v in row])
            return out.getvalue(), "xlsx rendered as csv text per sheet"
        if suffix == ".pdf":
            from pypdf import PdfReader

            pages = PdfReader(io.BytesIO(raw)).pages
            text = "\n".join(
                f"[page {i}]\n{page.extract_text() or ''}"
                for i, page in enumerate(pages, start=1)
            )
            return text, "pdf rendered as extracted text per page"
    except Exception as e:
        print(f"[outputs] preview failed for {path.name}: {e}", flush=True)
    return None

def _read_output(path: Path) -> tuple[bytes, str | None, str | None]:
    """Read one deliverable and return (raw bytes, judge-visible text or None, render note).
    Renderable formats go to _render_binary FIRST: a PDF can be pure ASCII, and the utf-8
    decode would hand the judge raw PDF syntax instead of per-page text."""
    raw = path.read_bytes()
    if rendering := _render_binary(path, raw):
        return raw, rendering[0], rendering[1]
    try:
        return raw, raw.decode("utf-8"), None
    except UnicodeDecodeError:
        return raw, None, None

def _budget_outputs(lengths: list[int]) -> list[int]:
    """Split OUTPUT_TOTAL_MAX_CHARS over the files by water-filling on their judge-visible
    text length: smallest first, each taking at most an equal share of what is left, so the
    slack the small files leave is redistributed to the large ones. A file with no
    judge-visible text takes 0 and its share goes to the others.

    First-come allocation in alphabetical order handed the whole budget to whichever file
    sorted earliest. On bub2-058 two raw scratch dumps took 500k of 600k and the primary
    deliverable, unique_mallorca_inventory.csv, was cut at the boundary -- the graded answer
    truncated so scratch could be shown whole.

    """
    limits = [0] * len(lengths)
    remaining = OUTPUT_TOTAL_MAX_CHARS
    for taken, i in enumerate(sorted(range(len(lengths)), key=lambda j: lengths[j])):
        share = remaining // (len(lengths) - taken)
        limits[i] = min(lengths[i], share)
        remaining -= limits[i]
    return limits

def _clip(text: str, limit: int) -> str:
    """Middle-out, never silent -- the same reason the judge truncates that way. A head-only
    cut drops the END of a tool result, where totals, the last rows and the final state live;
    the judge then saw a deliverable claim with its supporting evidence missing and read that
    as fabrication."""
    if len(text) <= limit:
        return text
    template = "...[{} characters omitted from the middle]..."
    budget = limit - len(template.format(f"{len(text):,}"))
    if budget <= 0:
        return text[:limit]
    marker = template.format(f"{len(text) - budget:,}")
    head = budget // 3
    return text[:head] + marker + text[-(budget - head):]

def collect(directory: Path) -> dict:
    # Skip SDK-owned observations and package installs. Do not follow symlinks.
    paths = sorted(
        p for p in directory.rglob("*")
        if p.is_file() and not p.is_symlink()
        and not any(part in {".browser-use", "node_modules", ".git", "screenshots"}
                    for part in p.relative_to(directory).parts)
    )
    rendered = [_read_output(p) for p in paths[:OUTPUT_MAX_FILES]]
    limits = _budget_outputs([len(text or "") for _, text, _ in rendered])
    files = []
    for path, (raw, text, kind), limit in zip(paths, rendered, limits):
        entry = {"name": str(path.relative_to(directory)), "size": len(raw), "text": None}
        if kind:
            entry["rendered"] = kind
        if text is not None:
            entry.update(text=_clip(text, limit), clipped=len(text) > limit)
        files.append(entry)
    return {
        "output_files": files,
        "staged_outputs": [str(p.relative_to(directory.parent)) for p in paths],
    }


if __name__ == "__main__":
    Path(sys.argv[2]).write_text(json.dumps(collect(Path(sys.argv[1]))) + "\n")
